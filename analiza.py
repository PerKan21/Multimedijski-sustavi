"""
Identifikacija govornika i obrada snimki
=========================================
Identifikacija VAD segmenata, clustering za broj govornika,
agregacija rezultata i spremanje u datoteku.
"""

import os
import numpy as np
from datetime import datetime
from sklearn.cluster import AgglomerativeClustering

from predobrada import predobradi_signal, vad_segmentacija, normaliziraj_segment
from model import izvuci_embedding_sa_segmentacijom, izvuci_embedding_iz_signala


# ================================================================
# IDENTIFIKACIJA - dva praga (prijedlog profesora)
# ================================================================
def identificiraj(embedding_ulaz: np.ndarray, baza: dict,
                  prag_donji: float, prag_gornji: float) -> tuple:
    """
    Usporeduje embedding s bazom i vraca (student, distanca, status).
    status: SIGURAN / NESIGURAN / ULJEZ
    """
    najbolji_student    = None
    najmanja_udaljenost = float("inf")
    emb = embedding_ulaz / np.linalg.norm(embedding_ulaz)

    for ime, referentni in baza.items():
        min_dist = min(float(1 - np.dot(emb, ref)) for ref in referentni)
        if min_dist < najmanja_udaljenost:
            najmanja_udaljenost = min_dist
            najbolji_student    = ime

    if najmanja_udaljenost <= prag_donji:
        return najbolji_student, najmanja_udaljenost, "SIGURAN"
    elif najmanja_udaljenost <= prag_gornji:
        return najbolji_student, najmanja_udaljenost, "NESIGURAN"
    else:
        return None, najmanja_udaljenost, "ULJEZ"


# ================================================================
# CLUSTERING - procjena broja govornika
# ================================================================
def procijeni_broj_govornika(embeddinzi: list, prag: float = 0.08) -> int:
    """
    Agglomerative clustering embeddinga VAD segmenata —
    procjenjuje koliko razlicitih govornika ima na snimci.
    Koristi 'complete' linkage — osjetljivije na razlike između govornika.
    """
    if len(embeddinzi) < 2:
        return len(embeddinzi)
    X = np.array(embeddinzi)
    clustering = AgglomerativeClustering(
        n_clusters=None, distance_threshold=prag,
        metric="cosine", linkage="complete"
    )
    clustering.fit(X)
    return len(set(clustering.labels_))


# ================================================================
# OBRADA JEDNE SNIMKE
# ================================================================
def obradi_snimku(putanja: str, baza: dict,
                  prag_donji: float, prag_gornji: float,
                  sr: int = 16000,
                  trajanje: float = 1.5,
                  preklapanje: float = 0.3,
                  prop_decrease: float = 0.0,
                  vad_top_db: float = 25,
                  vad_min_duljina: float = 0.3,
                  vad_spajanje: float = 0.15,
                  clustering_prag: float = 0.25) -> tuple:
    """
    Vraca (prepoznati, segmenti, uljezi_seg, n_govornika).
    segmenti = lista (poc, kraj, student, dist, status)
    """
    from predobrada import ucitaj_sirovi_signal, ukloni_sum, normaliziraj_segment

    # Isti preprocessing pipeline kao za bazu:
    # resample → denoising → VAD → normalizacija po segmentu
    signal, vad_seg = predobradi_signal(
        putanja, sr, prop_decrease,
        vad_top_db, vad_min_duljina, vad_spajanje
    )

    if not vad_seg:
        return [], [], 0, 0

    svi_segmenti    = []
    embeddinzi_svih = []
    uljezi_seg      = 0

    offset = 0
    for poc, kraj in vad_seg:
        duljina = int((kraj - poc) * sr)
        seg = signal[offset:offset + duljina]
        offset += duljina
        if len(seg) == 0:
            continue
        emb = izvuci_embedding_sa_segmentacijom(seg, sr, trajanje, preklapanje)
        embeddinzi_svih.append(emb)
        student, dist, status = identificiraj(emb, baza, prag_donji, prag_gornji)
        svi_segmenti.append((poc, kraj, student, dist, status))

    # Zadrzavamo samo najbolji segment po studentu
    najbolji_po_studentu = {}
    for poc, kraj, student, dist, status in svi_segmenti:
        if student is None:
            uljezi_seg += 1
            continue
        if student not in najbolji_po_studentu or dist < najbolji_po_studentu[student][2]:
            najbolji_po_studentu[student] = (poc, kraj, dist, status)

    filtrirani = []
    prepoznati = []
    vidjeni    = set()  # Za deduplikaciju
    for poc, kraj, student, dist, status in svi_segmenti:
        if student is None:
            filtrirani.append((poc, kraj, None, dist, status))
        else:
            best = najbolji_po_studentu.get(student)
            if best and best[0] == poc and best[1] == kraj:
                filtrirani.append((poc, kraj, student, dist, status))
                if status in ("SIGURAN", "NESIGURAN"):
                    if student not in vidjeni:
                        prepoznati.append(student)
                        vidjeni.add(student)

    # Broj govornika = broj jedinstveno identificiranih + ima li uljeza
    n_govornika = len(set(prepoznati)) + (1 if uljezi_seg > 0 else 0)

    return prepoznati, filtrirani, uljezi_seg, n_govornika


# ================================================================
# POMOCNA FUNKCIJA
# ================================================================
def fmt_s(sekunde: float) -> str:
    m = int(sekunde // 60)
    s = sekunde % 60
    return f"{m:02d}:{s:05.2f}"


# ================================================================
# SPREMANJE REZULTATA
# ================================================================
def spremi_rezultate(prisutnost: dict, svi_rezultati: dict,
                     putanja: str = "prisutnost.txt",
                     timestamp: str = None):
    datum_vrijeme = timestamp or datetime.now().strftime("%d.%m.%Y. u %H:%M:%S")
    prisutni  = [ime for ime, p in prisutnost.items() if p]
    nedostaju = [ime for ime, p in prisutnost.items() if not p]
    sep = "=" * 55 + "\n"

    with open(putanja, "w", encoding="utf-8") as f:
        f.write(f"Analiza provedena: {datum_vrijeme}\n")
        f.write(sep)
        for naziv, (prepoznati, segmenti, _, n_gov) in svi_rezultati.items():
            f.write(f"Snimka: {naziv}\n")
            f.write("\n")
            f.write(f"  Procijenjeni broj govornika: {n_gov}\n")
            f.write("\n")
            for poc, kraj, student, dist, status in segmenti:
                ime    = student if student else "NEPOZNAT"
                oznaka = "?" if status == "NESIGURAN" else ("+" if student else "!")
                f.write(f"  [{oznaka}] {fmt_s(poc)} - {fmt_s(kraj)}"
                        f"  {ime} (dist={dist:.4f}, {status})\n")
            f.write("\n")
            f.write(f"  Prepoznati: {', '.join(prepoznati) if prepoznati else 'nitko'}\n")
            f.write("\n")

            # Popis prisutnosti po snimci
            prepoznati_set = set(prepoznati)
            nedostaju_sn   = sorted(set(prisutnost.keys()) - prepoznati_set)
            f.write(f"  Prisutni na snimci ({len(prepoznati_set)}/{len(prisutnost)}):\n")
            for ime in prepoznati:
                f.write(f"    + {ime}\n")
            f.write("\n")
            if nedostaju_sn:
                f.write(f"  Nisu na snimci:\n")
                for ime in nedostaju_sn:
                    f.write(f"    - {ime}\n")
            f.write("\n")
            f.write(sep)




# ================================================================
# EXCEL EXPORT
# ================================================================
def spremi_excel(prisutnost: dict, svi_rezultati: dict,
                 putanja: str = "prisutnost.xlsx",
                 timestamp: str = None):
    """
    Sprema rezultate u Excel tablicu s bojama.
    List 1: Matrica studenti x snimke (zeleno = prisutan, crveno = odsutan)
    List 2: Detalji segmenata po snimci
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  GREŠKA: Instaliraj openpyxl: pip install openpyxl")
        return

    from datetime import datetime
    datum_vrijeme = timestamp or datetime.now().strftime("%d.%m.%Y. u %H:%M:%S")

    wb = Workbook()

    # ================================================================
    # PALETA BOJA (konzistentno s GUI-jem)
    # ================================================================
    fill_success   = PatternFill("solid", fgColor="3d7a42")
    fill_warn_3    = PatternFill("solid", fgColor="2a6a2a")   # 66-99%
    fill_warn_2    = PatternFill("solid", fgColor="6a6a10")   # 33-66%
    fill_warn_1    = PatternFill("solid", fgColor="7a4a1a")   # 1-33%
    fill_danger    = PatternFill("solid", fgColor="7a3d3d")
    fill_zaglavlje = PatternFill("solid", fgColor="1e1e1e")
    fill_header    = PatternFill("solid", fgColor="2a2a2a")
    fill_sub       = PatternFill("solid", fgColor="333333")
    fill_nesiguran = PatternFill("solid", fgColor="7a5c2a")

    font_naslov      = Font(color="e8e8e8", bold=True,  name="Calibri", size=14)
    font_bijeli_bold = Font(color="e8e8e8", bold=True,  name="Calibri", size=11)
    font_bijeli      = Font(color="e8e8e8", bold=False, name="Calibri", size=10)
    font_success     = Font(color="7ecf85", bold=True,  name="Calibri", size=11)
    font_warn_3      = Font(color="7ecf85", bold=True,  name="Calibri", size=11)
    font_warn_2      = Font(color="d4d44a", bold=True,  name="Calibri", size=11)
    font_warn_1      = Font(color="e8943a", bold=True,  name="Calibri", size=11)

    def boja_prisutnosti(n, ukupno):
        """Vraća (fill, font) ovisno o omjeru n/ukupno — konzistentno s GUI-jem."""
        if ukupno == 0 or n == 0:
            return fill_danger, font_danger
        omjer = n / ukupno
        if omjer == 1.0:
            return fill_success, font_success
        elif omjer >= 0.66:
            return fill_warn_3, font_warn_3
        elif omjer >= 0.33:
            return fill_warn_2, font_warn_2
        else:
            return fill_warn_1, font_warn_1

    font_naslov      = Font(color="e8e8e8", bold=True,  name="Calibri", size=14)
    font_bijeli_bold = Font(color="e8e8e8", bold=True,  name="Calibri", size=11)
    font_bijeli      = Font(color="e8e8e8", bold=False, name="Calibri", size=10)
    font_success     = Font(color="7ecf85", bold=True,  name="Calibri", size=11)
    font_danger      = Font(color="cf7e7e", bold=True,  name="Calibri", size=11)
    font_mute        = Font(color="888888",              name="Calibri", size=10)
    font_nesig       = Font(color="f0c080",              name="Calibri", size=10)
    font_warn        = Font(color="f0a030", bold=True,  name="Calibri", size=11)

    tanki  = Side(style="thin",   color="555555")
    srednji = Side(style="medium", color="888888")
    rub    = Border(left=tanki,  right=tanki,  top=tanki,    bottom=tanki)
    rub_jaci = Border(left=srednji, right=srednji, top=srednji, bottom=srednji)

    cen  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lij  = Alignment(horizontal="left",   vertical="center")
    lij_w = Alignment(horizontal="left",  vertical="center", wrap_text=True)

    studenti = sorted(prisutnost.keys())
    snimke   = list(svi_rezultati.keys())
    n_cols   = len(snimke) + 2  # +1 student, +1 ukupno

    # ================================================================
    # LIST 1: Matrica prisutnosti
    # ================================================================
    ws1       = wb.active
    ws1.title = "Prisutnost po snimci"
    ws1.sheet_view.showGridLines = False

    # --- Naslov ---
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws1.cell(1, 1, value="Sustav za popisivanje studenata")
    c.fill = fill_zaglavlje; c.font = font_naslov; c.alignment = cen
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws1.cell(2, 1, value=f"Analiza provedena: {datum_vrijeme}  |  Studenata: {len(studenti)}  |  Snimki: {len(snimke)}")
    c.fill = fill_zaglavlje; c.font = font_mute; c.alignment = cen
    ws1.row_dimensions[2].height = 20

    # --- Prazni red ---
    ws1.row_dimensions[3].height = 8
    for col in range(1, n_cols + 1):
        ws1.cell(3, col).fill = fill_zaglavlje

    # --- Zaglavlje stupaca ---
    c = ws1.cell(4, 1, value="Student")
    c.fill = fill_header; c.font = font_bijeli_bold; c.alignment = cen; c.border = rub_jaci
    ws1.column_dimensions["A"].width = 18

    for col, naziv in enumerate(snimke, 2):
        # Puno ime snimke — wrapa u ćeliju
        c = ws1.cell(4, col, value=naziv)
        c.fill = fill_header; c.font = font_bijeli_bold
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = rub_jaci
        ws1.column_dimensions[get_column_letter(col)].width = 16

    # Zadnji stupac — ukupno
    col_uk = len(snimke) + 2
    c = ws1.cell(4, col_uk, value="Ukupno\nprisutan")
    c.fill = fill_sub; c.font = font_bijeli_bold; c.alignment = cen; c.border = rub_jaci
    ws1.column_dimensions[get_column_letter(col_uk)].width = 12
    ws1.row_dimensions[4].height = 52

    # --- Redovi studenata ---
    for row, student in enumerate(studenti, 5):
        c = ws1.cell(row, 1, value=student)
        c.fill = fill_sub; c.font = font_bijeli_bold; c.alignment = lij; c.border = rub

        ukupno_prisutan = 0
        for col, naziv in enumerate(snimke, 2):
            prepoznati_sn, _, _, _ = svi_rezultati[naziv]
            prisutan = student in prepoznati_sn
            if prisutan:
                ukupno_prisutan += 1
            fill_s, font_s = boja_prisutnosti(1 if prisutan else 0, 1)
            tekst  = "✓" if prisutan else "✗"

            c = ws1.cell(row, col, value=tekst)
            c.fill = fill_s; c.font = font_s; c.alignment = cen; c.border = rub

        # Stupac ukupno — dinamička boja po omjeru
        fill_uk, font_uk = boja_prisutnosti(ukupno_prisutan, len(snimke))
        c = ws1.cell(row, col_uk, value=f"{ukupno_prisutan}/{len(snimke)}")
        c.fill = fill_uk; c.font = font_uk; c.alignment = cen; c.border = rub
        ws1.row_dimensions[row].height = 22

    # --- Red ukupno po snimci ---
    row_uk = len(studenti) + 5
    ws1.row_dimensions[row_uk].height = 22
    c = ws1.cell(row_uk, 1, value="Ukupno na snimci")
    c.fill = fill_sub; c.font = font_bijeli_bold; c.alignment = lij; c.border = rub_jaci

    for col, naziv in enumerate(snimke, 2):
        prepoznati_sn, _, _, _ = svi_rezultati[naziv]
        n = len(prepoznati_sn)
        fill_s, font_s = boja_prisutnosti(n, len(studenti))
        c = ws1.cell(row_uk, col, value=f"{n}/{len(studenti)}")
        c.fill = fill_s; c.font = font_s; c.alignment = cen; c.border = rub_jaci

    ws1.cell(row_uk, col_uk).fill = fill_zaglavlje
    ws1.cell(row_uk, col_uk).border = rub_jaci

    ws1.freeze_panes = "B5"

    # ================================================================
    # LIST 2: Detalji po snimci
    # ================================================================
    ws2       = wb.create_sheet("Detalji po snimci")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    c = ws2.cell(1, 1, value="Detalji analize po snimci")
    c.fill = fill_zaglavlje; c.font = font_naslov; c.alignment = cen
    ws2.row_dimensions[1].height = 36

    ws2.merge_cells("A2:G2")
    c = ws2.cell(2, 1, value=f"Analiza provedena: {datum_vrijeme}")
    c.fill = fill_zaglavlje; c.font = font_mute; c.alignment = cen
    ws2.row_dimensions[2].height = 20

    ws2.row_dimensions[3].height = 8
    for col in range(1, 8):
        ws2.cell(3, col).fill = fill_zaglavlje

    zaglavlja2 = ["Snimka", "Početak", "Kraj", "Trajanje (s)", "Govornik", "Distanca", "Status"]
    for col, zag in enumerate(zaglavlja2, 1):
        c = ws2.cell(4, col, value=zag)
        c.fill = fill_header; c.font = font_bijeli_bold
        c.alignment = cen; c.border = rub_jaci
    ws2.row_dimensions[4].height = 28

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 11
    ws2.column_dimensions["C"].width = 11
    ws2.column_dimensions["D"].width = 13
    ws2.column_dimensions["E"].width = 16
    ws2.column_dimensions["F"].width = 12
    ws2.column_dimensions["G"].width = 14

    trenutni_red = 5
    for naziv, (prepoznati, segmenti, _, n_gov) in svi_rezultati.items():
        for poc, kraj, student, dist, status in segmenti:
            ime_studenta = student if student else "NEPOZNAT"

            if status == "SIGURAN":
                fill_s = fill_success; font_s = font_success
            elif status == "NESIGURAN":
                fill_s = fill_nesiguran; font_s = font_nesig
            else:
                fill_s = fill_danger; font_s = font_danger

            vrijednosti = [naziv, fmt_s(poc), fmt_s(kraj),
                           f"{(kraj-poc):.1f}", ime_studenta,
                           f"{dist:.4f}", status]

            for col, vrijednost in enumerate(vrijednosti, 1):
                c = ws2.cell(trenutni_red, col, value=vrijednost)
                c.fill = fill_s; c.font = font_s
                c.alignment = cen if col != 1 else lij
                c.border = rub
            ws2.row_dimensions[trenutni_red].height = 18
            trenutni_red += 1

    ws2.freeze_panes = "A5"

    wb.save(putanja)
    print(f"  Excel tablica spremljena: {putanja}")